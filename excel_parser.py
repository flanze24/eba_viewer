"""
excel_parser.py
Extracts cell data, styles and structure from .xlsx files using openpyxl.
"""

from __future__ import annotations
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class CellStyle:
    bg_color: str | None = None
    fg_color: str | None = None
    bold: bool = False
    italic: bool = False
    font_size: float | None = None   # intentionally not stored (see _extract_style)
    font_name: str | None = None     # intentionally not stored
    h_align: str = "left"
    v_align: str = "top"
    wrap_text: bool = False
    border_top: str | None = None
    border_bottom: str | None = None
    border_left: str | None = None
    border_right: str | None = None
    number_format: str | None = None


@dataclass
class CellData:
    value: Any = None
    display_value: str = ""
    style: CellStyle = field(default_factory=CellStyle)
    rowspan: int = 1
    colspan: int = 1
    is_merged_hidden: bool = False
    coordinate: str | None = None   # e.g. "C 01.00,0010,0020"


@dataclass
class SheetData:
    name: str
    rows: list[list[CellData]] = field(default_factory=list)
    col_widths: list[float] = field(default_factory=list)
    row_heights: list[float] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_THEME_COLORS = [
    "FFFFFF", "000000", "E7E6E6", "44546A", "4472C4",
    "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47",
]

# Regex: exactly 4 decimal digits (EBA DPM codes like 0010, 0020, ...)
_FOUR_DIGIT = re.compile(r'^\d{4}$')


def _apply_tint(hex_color: str, tint: float) -> str:
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        if tint < 0:
            factor = 1 + tint
            r, g, b = int(r * factor), int(g * factor), int(b * factor)
        else:
            r = int(r + (255 - r) * tint)
            g = int(g + (255 - g) * tint)
            b = int(b + (255 - b) * tint)
        return f"{r:02X}{g:02X}{b:02X}"
    except Exception:
        return hex_color


def _resolve_color(color_obj, theme_colors: list[str], ignore_alpha: bool = False) -> str | None:
    """Convert an openpyxl Color object to a plain 6-char hex string.

    ignore_alpha=True: used for PatternFill backgrounds.  Excel stores solid
    fills with alpha=00 in many cases but still renders the colour — we must
    do the same and read only the RGB bytes.
    """
    if color_obj is None:
        return None
    try:
        t = color_obj.type
        if t == "rgb":
            raw = color_obj.rgb
            if not raw:
                return None
            alpha = raw[:2].upper()
            hex6  = raw[-6:].upper()
            if raw.upper() in ("00000000", "FF000000"):
                return None
            if ignore_alpha:
                # Black and white are indistinguishable from "no fill" on screen
                if hex6 in ("000000", "FFFFFF"):
                    return None
                return hex6
            else:
                if alpha == "00":
                    return None
                return hex6
        elif t == "theme":
            idx = color_obj.theme
            if 0 <= idx < len(theme_colors):
                base = theme_colors[idx]
                tint = getattr(color_obj, "tint", 0.0) or 0.0
                if tint != 0.0:
                    base = _apply_tint(base, tint)
                return base
        elif t == "indexed":
            _INDEXED = {
                 0:"000000", 1:"FFFFFF", 2:"FF0000", 3:"00FF00",
                 4:"0000FF", 5:"FFFF00", 6:"FF00FF", 7:"00FFFF",
                 8:"000000", 9:"FFFFFF",16:"800000",17:"008000",
                18:"000080",19:"808000",20:"800080",21:"008080",
                22:"C0C0C0",23:"808080",24:"9999FF",25:"993366",
                26:"FFFFCC",27:"CCFFFF",28:"660066",29:"FF8080",
                30:"0066CC",31:"CCCCFF",40:"00CCFF",41:"CCFFFF",
                42:"CCFFCC",43:"FFFF99",44:"99CCFF",45:"FF99CC",
                46:"CC99FF",47:"FFCC99",48:"3366FF",49:"33CCCC",
                50:"99CC00",51:"FFCC00",52:"FF9900",53:"FF6600",
                54:"666699",55:"969696",56:"003366",57:"339966",
                63:"333333",64:"000000",
            }
            idx_val = getattr(color_obj, "indexed", None)
            if idx_val is not None:
                try:
                    return _INDEXED.get(int(idx_val))
                except Exception:
                    pass
    except Exception:
        pass
    return None


def _border_style(side) -> str | None:
    if side and side.border_style and side.border_style != "none":
        styles = {
            "thin": "1px solid", "medium": "2px solid", "thick": "3px solid",
            "dashed": "1px dashed", "dotted": "1px dotted",
            "double": "3px double", "hair": "1px solid",
        }
        css = styles.get(side.border_style, "1px solid")
        color = None
        if side.color:
            try:
                raw = side.color.rgb
                if raw:
                    color = f"#{raw[-6:]}"
            except Exception:
                pass
        color = color or "#BDBDBD"
        return f"{css} {color}"
    return None


def _format_value(value: Any, number_format: str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        if number_format and "%" in number_format:
            return f"{value * 100:.2f}%"
        if isinstance(value, float):
            if abs(value) < 1e9:
                s = f"{value:,.4f}".rstrip("0").rstrip(".")
                return s
        return str(value)
    return str(value)



def _is_near_white(hex6: str, threshold: int = 248) -> bool:
    """Return True if all RGB components are >= threshold.

    Covers pure white (FFFFFF), near-white theme tints (FCFCFC, FAFAFA, …)
    and any other visually-empty light fill that should not block coordinates.
    """
    try:
        r = int(hex6[0:2], 16)
        g = int(hex6[2:4], 16)
        b = int(hex6[4:6], 16)
        return r >= threshold and g >= threshold and b >= threshold
    except Exception:
        return False

def _cell_is_input(cell_data: CellData) -> bool:
    """Return True when a cell should receive a coordinate label.

    A cell is an input cell when:
    - it has no background colour fill, AND
    - it has no visible text content
    Merged-hidden cells are never labelled (they have no visual presence).
    """
    if cell_data.is_merged_hidden:
        return False
    has_bg   = bool(cell_data.style.bg_color)
    has_text = bool(cell_data.display_value.strip())
    return (not has_bg) and (not has_text)


# ---------------------------------------------------------------------------
# Coordinate builder
# ---------------------------------------------------------------------------

def _build_coordinates(sheet: SheetData) -> None:
    """
    Assign coordinate strings to input cells.

    Coordinate format: "<sheet_name>,<col_code>,<row_code>"
    where col_code and row_code are 4-digit strings found in the header
    row/column of the sheet.

    Discovery rules:
    - Column codes: scan the first 15 rows for a row whose cells contain
      at least 1 four-digit value.  That row index is the "column-header row".
      The column position of each four-digit value is its column index.
    - Row codes: scan the first column for four-digit values.  The row
      position of each four-digit value is its row index.
    - Only cells that are BOTH in a row that has a row-code AND in a column
      that has a column-code receive a coordinate.
    - Cells that already have a bg_color or have text content are skipped.
    """
    rows = sheet.rows
    if not rows:
        return

    n_rows = len(rows)
    n_cols = len(rows[0]) if rows else 0

    # ── 1. Find the column-header row ────────────────────────────────────────
    # Scan first 15 rows; pick the first one with ≥1 four-digit codes.
    col_code_row: int | None = None   # 0-based index into rows
    col_index_to_code: dict[int, str] = {}  # col 0-based → "0010"

    for ri in range(min(15, n_rows)):
        codes_in_row: dict[int, str] = {}
        for ci, cell in enumerate(rows[ri]):
            v = cell.display_value.strip()
            if _FOUR_DIGIT.match(v):
                codes_in_row[ci] = v
        if len(codes_in_row) >= 1:
            col_code_row = ri
            col_index_to_code = codes_in_row
            break

    # ── 2. Find row codes – scan first 5 columns ─────────────────────────────
    # Pick the column (among the first 5) with the most four-digit values;
    # that column carries the row codes.
    row_index_to_code: dict[int, str] = {}  # row 0-based → "0010"

    best_col: int = 0
    best_count: int = 0
    for ci in range(min(5, n_cols)):
        count = sum(
            1 for ri in range(n_rows)
            if rows[ri] and ci < len(rows[ri])
            and _FOUR_DIGIT.match(rows[ri][ci].display_value.strip())
        )
        if count > best_count:
            best_count = count
            best_col = ci

    if best_count > 0:
        for ri in range(n_rows):
            if not rows[ri] or best_col >= len(rows[ri]):
                continue
            v = rows[ri][best_col].display_value.strip()
            if _FOUR_DIGIT.match(v):
                row_index_to_code[ri] = v

    if not col_index_to_code or not row_index_to_code:
        return   # sheet has no recognisable DPM coordinate system

    # ── 3. Assign coordinates ────────────────────────────────────────────────
    sheet_name = sheet.name
    for ri, row_cells in enumerate(rows):
        row_code = row_index_to_code.get(ri)
        if row_code is None:
            continue  # rows without a row-code get no coordinate
        for ci, cell in enumerate(row_cells):
            col_code = col_index_to_code.get(ci)
            if col_code is None:
                continue
            if _cell_is_input(cell):
                cell.coordinate = f"{sheet_name}_{row_code}_{col_code}"


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_workbook(path: str | Path) -> dict[str, SheetData]:
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)

    theme_colors: list[str] = list(_THEME_COLORS)
    try:
        if wb.loaded_theme:
            import xml.etree.ElementTree as ET
            ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
            root = ET.fromstring(wb.loaded_theme)
            palette = []
            for tag in ["lt1", "dk1", "lt2", "dk2"]:
                el = root.find(f".//a:{tag}//a:srgbClr", ns)
                if el is not None:
                    palette.append(el.get("val", "000000").upper())
            for el in root.findall(".//a:accent*//a:srgbClr", ns):
                palette.append(el.get("val", "000000").upper())
            if len(palette) >= 4:
                theme_colors = palette + _THEME_COLORS[len(palette):]
    except Exception:
        pass

    result: dict[str, SheetData] = {}
    for ws in wb.worksheets:
        state = getattr(ws, "sheet_state", "visible") or "visible"
        if state != "visible":
            continue
        sheet_data = _parse_sheet(ws, theme_colors)
        _build_coordinates(sheet_data)
        result[ws.title] = sheet_data

    wb.close()
    return result


def _parse_sheet(ws, theme_colors: list[str]) -> SheetData:
    merge_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    hidden_cells: set[tuple[int, int]] = set()
    for merged_range in ws.merged_cells.ranges:
        min_r, min_c = merged_range.min_row, merged_range.min_col
        max_r, max_c = merged_range.max_row, merged_range.max_col
        rowspan = max_r - min_r + 1
        colspan = max_c - min_c + 1
        merge_map[(min_r, min_c)] = (rowspan, colspan, max_r, max_c)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    hidden_cells.add((r, c))

    col_widths: list[float] = []
    for i in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(i)
        cd = ws.column_dimensions.get(letter)
        width = (cd.width if cd and cd.width else 8.43)
        col_widths.append(max(width, 2.0))

    row_heights: list[float] = []
    for i in range(1, (ws.max_row or 1) + 1):
        rd = ws.row_dimensions.get(i)
        height = (rd.height if rd and rd.height else 15.0)
        row_heights.append(max(height, 10.0))

    raw_rows: list[list[CellData]] = []
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    for r_idx in range(1, max_row + 1):
        row_cells: list[CellData] = []
        for c_idx in range(1, max_col + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            is_hidden = (r_idx, c_idx) in hidden_cells
            style = _extract_style(cell, theme_colors)
            value = cell.value
            display = _format_value(value, style.number_format)
            if (r_idx, c_idx) in merge_map:
                rs, cs, _, _ = merge_map[(r_idx, c_idx)]
            else:
                rs, cs = 1, 1
            row_cells.append(CellData(
                value=value, display_value=display, style=style,
                rowspan=rs, colspan=cs, is_merged_hidden=is_hidden,
            ))
        raw_rows.append(row_cells)

    def row_is_empty(row):
        return all(c.display_value == "" and not c.is_merged_hidden for c in row)

    def col_is_empty(j):
        return all(
            raw_rows[r][j].display_value == "" and not raw_rows[r][j].is_merged_hidden
            for r in range(len(raw_rows))
        )

    kept_rows = [i for i, row in enumerate(raw_rows) if not row_is_empty(row)]
    kept_cols = [j for j in range(max_col) if not col_is_empty(j)]

    filtered_rows = [[raw_rows[r][c] for c in kept_cols] for r in kept_rows]
    filtered_widths  = [col_widths[c]  for c in kept_cols]
    filtered_heights = [row_heights[r] for r in kept_rows]

    return SheetData(
        name=ws.title,
        rows=filtered_rows,
        col_widths=filtered_widths,
        row_heights=filtered_heights,
    )


def _extract_style(cell, theme_colors: list[str]) -> CellStyle:
    style = CellStyle()

    try:
        fill = cell.fill
        if fill and fill.fill_type not in (None, "none"):
            if isinstance(fill, PatternFill):
                if fill.fill_type == "solid":
                    color = _resolve_color(fill.fgColor, theme_colors, ignore_alpha=True)
                    if not color:
                        color = _resolve_color(fill.bgColor, theme_colors, ignore_alpha=True)
                    # Near-white fills (e.g. theme tints resolving to FCFCFC, FAFAFA …)
                    # are visually indistinguishable from no fill → treat as empty.
                    if color and _is_near_white(color):
                        color = None
                    style.bg_color = color
            else:
                color = _resolve_color(getattr(fill, "fgColor", None), theme_colors, ignore_alpha=True)
                if color and _is_near_white(color):
                    color = None
                style.bg_color = color
    except Exception:
        pass

    # Font – size and name intentionally NOT stored (unified in renderer)
    try:
        font = cell.font
        if font:
            style.bold   = bool(font.bold)
            style.italic = bool(font.italic)
            color = _resolve_color(font.color, theme_colors)
            style.fg_color = color
    except Exception:
        pass

    try:
        al = cell.alignment
        if al:
            h = al.horizontal or "left"
            h_map = {"general": "left", "centerContinuous": "center"}
            style.h_align  = h_map.get(h, h)
            style.v_align  = al.vertical or "top"
            style.wrap_text = bool(al.wrapText)
    except Exception:
        pass

    try:
        b = cell.border
        if b:
            style.border_top    = _border_style(b.top)
            style.border_bottom = _border_style(b.bottom)
            style.border_left   = _border_style(b.left)
            style.border_right  = _border_style(b.right)
    except Exception:
        pass

    try:
        style.number_format = cell.number_format
    except Exception:
        pass

    return style