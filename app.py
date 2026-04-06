"""
app.py – EBA ITS DPM Viewer (Streamlit)
========================================
Lädt eine Excel-Datei automatisch, stellt alle Tabellenblätter originalgetreu dar
und bietet eine Navigation über das "Index"-Blatt mit automatischer Gruppierung.

Konfiguration:
  Setze EXCEL_PATH auf den vollständigen Pfad zur Excel-Datei.
"""

from __future__ import annotations
import csv
import os
import sys
from pathlib import Path
from dataclasses import dataclass, field

import streamlit as st

# ── Konfiguration ──────────────────────────────────────────────────────────────
# Pfad zur Excel-Datei – kann auch als Umgebungsvariable gesetzt werden:
#   export EBA_EXCEL_PATH="/pfad/zur/datei.xlsx"
EXCEL_PATH = os.environ.get(
    "EBA_EXCEL_PATH",
    str(Path(__file__).parent / "data" / "C_2024_8389_F1_ANNEX_DE_V1_P1_3682615.XLSX"),
)
INDEX_SHEET = "Index"   # Name des Index-Blatts (case-sensitive)
APP_TITLE   = "EBA ITS DPM Viewer"
# ──────────────────────────────────────────────────────────────────────────────

# Add current dir to path so modules resolve correctly
sys.path.insert(0, str(Path(__file__).parent))

from excel_parser import parse_workbook, SheetData
from renderer import render_sheet_html


# ── Page setup ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title=APP_TITLE,
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Global CSS
st.markdown("""
<style>
/* App-wide font – single unified stack for all UI and table text */
html, body, [class*="css"], table, td, th {
    font-family: 'Segoe UI', 'Inter', 'Calibri', system-ui, sans-serif !important;
}

/* Single authoritative font-size for ALL table content – no exceptions */
table, table td, table th, table * {
    font-size: 10pt !important;
    line-height: 1.45 !important;
}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: #1B2A4A;
    color: #E8EFF8;
    min-width: 260px;
}
/* Group header labels */
section[data-testid="stSidebar"] .sidebar-group-label {
    color: #7EB8F7;
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    padding: 10px 12px 2px 12px;
    margin-top: 4px;
    display: block;
}
section[data-testid="stSidebar"] .stButton button {
    width: 100%;
    text-align: left;
    background: transparent;
    color: #C8D8F0;
    border: none;
    border-left: 3px solid transparent;
    border-radius: 0;
    padding: 4px 12px 4px 16px;
    font-size: 0.82rem;
    margin: 0;
    transition: all 0.15s;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
section[data-testid="stSidebar"] .stButton button:hover {
    background: rgba(255,255,255,0.08);
    border-left-color: #4D9FEC;
    color: #ffffff;
}
section[data-testid="stSidebar"] .stButton button[kind="primary"] {
    background: rgba(77,159,236,0.18) !important;
    border-left-color: #4D9FEC !important;
    color: #ffffff !important;
    font-weight: 600;
}
section[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.12);
    margin: 6px 0;
}
section[data-testid="stSidebar"] h1 {
    color: #E8EFF8;
    font-size: 1.1rem;
}

/* ── Main header ── */
.eba-header {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 10px 0 16px 0;
    border-bottom: 2px solid #2B5FA8;
    margin-bottom: 16px;
}
.eba-header h1 {
    margin: 0;
    font-size: 1.5rem;
    color: #1B2A4A;
    font-weight: 700;
}
.eba-header .badge {
    background: #2B5FA8;
    color: white;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.5px;
}

/* Sheet title bar */
.sheet-title-bar {
    display: flex;
    align-items: center;
    justify-content: space-between;
    background: #F0F4FA;
    border: 1px solid #C5D5E8;
    border-radius: 6px;
    padding: 8px 16px;
    margin-bottom: 12px;
}
.sheet-title-bar h2 {
    margin: 0;
    font-size: 1.1rem;
    color: #1B2A4A;
}
.back-btn {
    background: #2B5FA8 !important;
    color: white !important;
    border: none !important;
    border-radius: 4px !important;
    padding: 4px 14px !important;
    font-size: 0.82rem !important;
    cursor: pointer !important;
    transition: background 0.15s !important;
}
.back-btn:hover {
    background: #1B4A8A !important;
}

/* Scrollable table wrapper */
.table-scroll-wrapper {
    overflow: auto;
    max-height: 78vh;
    border: 1px solid #D0D8E8;
    border-radius: 4px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.06);
}

/* Index cards */
.index-card {
    background: white;
    border: 1px solid #C5D5E8;
    border-left: 4px solid #2B5FA8;
    border-radius: 6px;
    padding: 10px 16px;
    margin-bottom: 8px;
    cursor: pointer;
    transition: box-shadow 0.15s, border-left-color 0.15s;
}
.index-card:hover {
    box-shadow: 0 2px 8px rgba(43,95,168,0.2);
    border-left-color: #4D9FEC;
}

/* Error box */
.error-box {
    background: #FFF0F0;
    border: 1px solid #E0A0A0;
    border-radius: 6px;
    padding: 16px;
    color: #8B0000;
}

/* Stat chips */
.stat-chip {
    display:inline-block;
    background:#E8F0FC;
    color:#1B2A4A;
    border-radius:12px;
    padding:2px 10px;
    font-size:0.78rem;
    margin-right:6px;
}
</style>
""", unsafe_allow_html=True)


# ── Session state ──────────────────────────────────────────────────────────────
if "current_sheet" not in st.session_state:
    st.session_state.current_sheet = INDEX_SHEET


# ── Index structure dataclasses ───────────────────────────────────────────────
@dataclass
class IndexEntry:
    short_name: str      # sheet name, e.g. "CA1"
    template_code: str   # e.g. "C 01.00"
    template_name: str   # long name
    number: str          # template number string

@dataclass
class IndexGroup:
    label: str
    entries: list[IndexEntry] = field(default_factory=list)


# ── Index parsing ─────────────────────────────────────────────────────────────
def _parse_index_structure(path: str) -> list[IndexGroup]:
    """
    Read the Index sheet and return a list of IndexGroups by scanning
    section headers and template rows.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return []

    if INDEX_SHEET not in wb.sheetnames:
        wb.close()
        return []

    ws        = wb[INDEX_SHEET]
    sheet_names = set(wb.sheetnames)

    SECTION_MAP = {
        "COREP":  "COREP",
        "FINREP": "FINREP",
        "IP LOS": "IP Losses",
        "LARGE":  "Large Exposures",
        "LEVERA": "Leverage Ratio",
        "LIQUID": "Liquidity",
        "AMM TE": "AMM",
        "TEMPLA": "G-SII",
        "IRRBB":  "IRRBB",
    }
    SECTION_DISPLAY = {
        "COREP":  "COREP",
        "FINREP": "FINREP · IFRS",
        "IP LOS": "IP Losses",
        "LARGE":  "Large Exposures",
        "LEVERA": "Leverage Ratio",
        "LIQUID": "Liquidity",
        "AMM TE": "AMM",
        "TEMPLA": "G-SII",
        "IRRBB":  "IRRBB",
        # Extra sections that appear as standalone annex blocks
        "IP_VERLUSTE": "IP-Verluste",
        "VERSCHULDUNG": "Verschuldungsquote",
    }

    # Keywords that signal a new top-level section in the annex blocks
    # (matched against *any* non-empty cell in col B/C/D of header-only rows)
    ANNEX_KEYWORDS = {
        "IP-VERLUSTE":        "IP_VERLUSTE",
        "IP VERLUSTE":        "IP_VERLUSTE",
        "IMMOBILIENBESICHER": "IP_VERLUSTE",
        "VERSCHULDUNG":       "VERSCHULDUNG",
    }

    groups: list[IndexGroup] = []
    current_section  = ""
    current_subgroup = ""
    finrep_count     = 0

    def _c(v) -> str:
        return str(v).strip() if v is not None else ""

    def _find_sheet(candidate: str) -> str | None:
        """Return the actual sheet name for *candidate*, or None.

        Handles:
        - Exact match
        - Case-insensitive match
        - The Index lists 'LR6.1' and 'LR6.2' as separate entries, but the
          workbook has a single sheet named 'LR6.1, LR6.2'.  We therefore
          also look for any sheet whose name *contains* the candidate token
          (after normalising non-breaking spaces and punctuation).
        """
        c = candidate.strip().replace("\xa0", " ")
        if c in sheet_names:
            return c
        c_lower = c.lower()
        for sn in sheet_names:
            if sn.strip().lower() == c_lower:
                return sn
        # Partial / combined-sheet fallback:
        # e.g. candidate="LR6.1" → sheet="LR6.1, LR6.2"
        for sn in sheet_names:
            sn_norm = sn.strip().lower()
            if c_lower in sn_norm:
                return sn
        return None

    for row in ws.iter_rows(min_row=4, max_row=500):
        c2 = _c(row[1].value)   # col B = number
        c3 = _c(row[2].value)   # col C = template code
        c4 = _c(row[3].value)   # col D = name
        c5 = _c(row[4].value)   # col E = short name / sheet name

        # ── Top-level section rows (no code, no sheet name) ──────────────
        if not c3 and not c5:
            # Standard COREP/FINREP sections (keyword in col B)
            matched = False
            for kw, _ in SECTION_MAP.items():
                if c2.upper().startswith(kw.upper()):
                    if kw == "FINREP":
                        finrep_count += 1
                    current_section  = kw
                    current_subgroup = ""
                    matched = True
                    break

            if not matched:
                # Annex blocks: keyword may appear in col B (long descriptive text)
                combined = f"{c2} {c4}".upper().replace("\xa0", " ")
                for kw, section_key in ANNEX_KEYWORDS.items():
                    if kw in combined:
                        current_section  = section_key
                        current_subgroup = ""
                        break
            continue

        # ── Sub-group rows (no code, has name) ───────────────────────────
        if not c3 and c4 and not c5:
            current_subgroup = c4.title()
            continue

        # ── Template row: find matching sheet ────────────────────────────
        target_sheet = _find_sheet(c5)

        if target_sheet is None:
            continue

        # Build group label
        if finrep_count >= 2 and current_section == "FINREP":
            sec_display = "FINREP · GAAP"
        else:
            sec_display = SECTION_DISPLAY.get(current_section, current_section or "Sonstige")

        # Include subgroup for sections that benefit from it
        if current_subgroup and current_section in ("COREP", "FINREP", "IRRBB", "LIQUID"):
            group_label = f"{sec_display} · {current_subgroup.title()}"
        else:
            group_label = sec_display

        if not groups or groups[-1].label != group_label:
            groups.append(IndexGroup(label=group_label))

        # Deduplicate: don't add the same sheet twice within one group
        existing_names = {e.short_name for e in groups[-1].entries}
        if target_sheet in existing_names:
            continue

        groups[-1].entries.append(IndexEntry(
            short_name=target_sheet,
            template_code=c3,
            template_name=c4,
            number=c2,
        ))

    wb.close()
    return groups


# ── Data loading (cached) ──────────────────────────────────────────────────────
@st.cache_resource(show_spinner="⏳ Lade Excel-Datei …")
def load_workbook(path: str):
    try:
        from export_coordinates import export_coordinates
        csv_path = Path(path).parent / "coordinates.csv"
        export_coordinates(path, csv_path)
        sheets = parse_workbook(path)
        _apply_annotations(sheets, csv_path)
        groups = _parse_index_structure(path)
        return sheets, groups
    except FileNotFoundError:
        return None, []
    except Exception as exc:
        st.error(f"Fehler beim Laden der Datei: {exc}")
        return None, []


def _apply_annotations(sheets: dict[str, SheetData], csv_path: Path) -> None:
    """Read annotations from coordinates.csv and attach them to matching cells.
    Supports both full coordinates (input cells) and label keys (row/col headers).
    """
    if not csv_path.exists():
        return
    annotations: dict[str, str] = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or "annotation" not in reader.fieldnames:
            return
        # Support both old ("coordinate") and new ("key") CSV format
        key_field = "key" if "key" in reader.fieldnames else "coordinate"
        for row in reader:
            note = row.get("annotation", "").strip()
            if note:
                annotations[row.get(key_field, "")] = note
    if not annotations:
        return
    for sheet in sheets.values():
        for row in sheet.rows:
            for cell in row:
                if cell.coordinate and cell.coordinate in annotations:
                    cell.annotation = annotations[cell.coordinate]
                lk = getattr(cell, "label_key", None)
                if lk and lk in annotations:
                    cell.annotation = annotations[lk]


# ── Navigation helper ──────────────────────────────────────────────────────────
def go_to(sheet_name: str) -> None:
    st.session_state.current_sheet = sheet_name


# ── Sidebar ────────────────────────────────────────────────────────────────────
def render_sidebar(sheets: dict[str, SheetData], groups: list[IndexGroup]) -> None:
    with st.sidebar:
        st.markdown("## 📊 EBA DPM Viewer")
        st.markdown("---")

        # Index / Home button
        is_index = st.session_state.current_sheet == INDEX_SHEET
        if st.button(
            "🏠 Index",
            key="nav_index",
            type="primary" if is_index else "secondary",
        ):
            go_to(INDEX_SHEET)

        if groups:
            # Grouped navigation built from the parsed Index structure
            nav_counter = 0
            for grp in groups:
                st.markdown(
                    f'<span class="sidebar-group-label">{grp.label}</span>',
                    unsafe_allow_html=True,
                )
                for entry in grp.entries:
                    name = entry.short_name
                    if name not in sheets:
                        continue
                    is_active = st.session_state.current_sheet == name
                    label = f"{entry.template_code}  {name}" if entry.template_code else name
                    nav_counter += 1
                    if st.button(
                        label,
                        key=f"nav_{nav_counter}_{name}",
                        type="primary" if is_active else "secondary",
                        help=entry.template_name,
                    ):
                        go_to(name)
        else:
            # Fallback: flat list
            st.markdown(
                '<span class="sidebar-group-label">Tabellenblätter</span>',
                unsafe_allow_html=True,
            )
            for name in sheets:
                if name == INDEX_SHEET:
                    continue
                is_active = st.session_state.current_sheet == name
                if st.button(name, key=f"nav_{name}",
                             type="primary" if is_active else "secondary"):
                    go_to(name)

        st.markdown("---")
        st.markdown(
            f"<small style='color:#8899BB'>Datei:<br><code style='font-size:0.7rem'>"
            f"{Path(EXCEL_PATH).name}</code></small>",
            unsafe_allow_html=True,
        )


# ── Main header ────────────────────────────────────────────────────────────────
def render_header() -> None:
    st.markdown(
        f"""
        <div class="eba-header">
            <span style="font-size:2rem">📊</span>
            <h1>{APP_TITLE}</h1>
            <span class="badge">EBA ITS DPM</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ── Index page ─────────────────────────────────────────────────────────────────
def render_index(sheets: dict[str, SheetData], groups: list[IndexGroup]) -> None:
    render_header()

    index_data = sheets.get(INDEX_SHEET)
    if index_data is None:
        st.warning(f'Kein Tabellenblatt "{INDEX_SHEET}" gefunden.')
        _render_fallback_index(sheets)
        return

    # Stat bar
    n_sheets = len(sheets) - 1   # exclude Index itself
    st.markdown(
        f'<span class="stat-chip">📄 {n_sheets} Tabellenblätter</span>'
        f'<span class="stat-chip">📁 {Path(EXCEL_PATH).name}</span>',
        unsafe_allow_html=True,
    )
    st.markdown("")

    # Render the Index table (navigation via sidebar/expanders below)
    html_table = render_sheet_html(index_data)

    st.markdown('<div class="table-scroll-wrapper">', unsafe_allow_html=True)
    st.markdown(html_table, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Grouped quick-navigation below the table ───────────────────────────
    if groups:
        st.markdown("---")
        st.markdown("### 🔗 Navigation nach Themenbereich")
        for grp in groups:
            valid = [e for e in grp.entries if e.short_name in sheets]
            if not valid:
                continue
            with st.expander(grp.label, expanded=False):
                cols = st.columns(3)
                for i, entry in enumerate(valid):
                    with cols[i % 3]:
                        label = f"{entry.template_code}  {entry.short_name}" if entry.template_code else entry.short_name
                        if st.button(
                            label,
                            key=f"idx_{grp.label}_{i}_{entry.short_name}",
                            help=entry.template_name,
                            use_container_width=True,
                        ):
                            go_to(entry.short_name)
    else:
        st.markdown("---")
        st.markdown("### 🔗 Direktlinks zu allen Tabellenblättern")
        cols = st.columns(3)
        for i, name in enumerate(sheets):
            if name == INDEX_SHEET:
                continue
            with cols[i % 3]:
                if st.button(f"📋 {name}", key=f"idx_link_{name}", use_container_width=True):
                    go_to(name)




def _render_fallback_index(sheets: dict[str, SheetData]) -> None:
    """Fallback wenn kein 'Index' Blatt vorhanden ist."""
    st.markdown("### Alle verfügbaren Tabellenblätter")
    cols = st.columns(3)
    for i, name in enumerate(sheets):
        with cols[i % 3]:
            if st.button(f"📋 {name}", key=f"fb_{name}", use_container_width=True):
                go_to(name)


# ── Sheet page ─────────────────────────────────────────────────────────────────
def render_sheet(sheet: SheetData, groups: list[IndexGroup]) -> None:
    # Title bar with back button
    col_back, col_title = st.columns([1, 6])
    with col_back:
        if st.button("⬅ Index", key="back_to_index", help="Zurück zur Index-Seite"):
            go_to(INDEX_SHEET)
    with col_title:
        # Breadcrumb: find which group this sheet belongs to
        breadcrumb = ""
        for grp in groups:
            for e in grp.entries:
                if e.short_name == sheet.name:
                    breadcrumb = (
                        f"<span style='color:#8899BB;font-size:0.8rem'>"
                        f"{grp.label} &rsaquo;</span> "
                    )
                    break

        n_rows = len(sheet.rows)
        n_cols = len(sheet.col_widths)
        st.markdown(
            f"""
            <div style="display:flex;align-items:center;gap:12px;margin-top:4px">
                {breadcrumb}
                <h2 style="margin:0;color:#1B2A4A;font-size:1.25rem">📋 {sheet.name}</h2>
                <span class="stat-chip">{n_rows} Zeilen</span>
                <span class="stat-chip">{n_cols} Spalten</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("")

    # Render table
    html_table = render_sheet_html(sheet)
    st.markdown('<div class="table-scroll-wrapper">', unsafe_allow_html=True)
    st.markdown(html_table, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


# ── Error page ─────────────────────────────────────────────────────────────────
def render_error_page() -> None:
    render_header()
    st.markdown(
        f"""
        <div class="error-box">
            <h3>⚠️ Excel-Datei nicht gefunden</h3>
            <p>Die Datei konnte unter folgendem Pfad nicht geladen werden:</p>
            <code>{EXCEL_PATH}</code>
            <hr>
            <p><strong>Lösungen:</strong></p>
            <ul>
                <li>Kopiere deine Excel-Datei nach <code>{Path(EXCEL_PATH).parent}</code>
                    und benenne sie <code>{Path(EXCEL_PATH).name}</code>.</li>
                <li>Oder setze die Umgebungsvariable:<br>
                    <code>export EBA_EXCEL_PATH="/vollständiger/pfad/zur/datei.xlsx"</code></li>
                <li>Oder passe <code>EXCEL_PATH</code> direkt in <code>app.py</code> an.</li>
            </ul>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Upload fallback
    st.markdown("---")
    st.markdown("### 📤 Oder Datei jetzt hochladen (temporär)")
    uploaded = st.file_uploader("Excel-Datei hochladen", type=["xlsx", "xls"])
    if uploaded:
        import tempfile
        tmp = tempfile.mktemp(suffix=".xlsx")
        with open(tmp, "wb") as f:
            f.write(uploaded.read())
        # Clear cache and reload with uploaded file
        load_workbook.clear()
        os.environ["EBA_EXCEL_PATH"] = tmp
        st.rerun()


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    sheets, groups = load_workbook(EXCEL_PATH)

    if sheets is None:
        render_error_page()
        return

    render_sidebar(sheets, groups)

    current = st.session_state.current_sheet

    # Guard: if stored sheet no longer exists, fall back to index
    if current not in sheets:
        current = INDEX_SHEET
        st.session_state.current_sheet = current

    if current == INDEX_SHEET or current not in sheets:
        render_index(sheets, groups)
    else:
        render_sheet(sheets[current], groups)


if __name__ == "__main__":
    main()